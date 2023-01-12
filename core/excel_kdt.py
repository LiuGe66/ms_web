# -*- coding: utf-8 -*-            
# Author:liu_ge
# @FileName: excel_data.py
# @Time : 2022/11/27 11:15
from pathlib import Path
from core.setting import settings
import allure
import pytest
from openpyxl.reader.excel import load_workbook
from core.kdt import KeyWord
from utils.logger_utils import *

sheet_name = []
step_id = []


def filter_empty(old_l):
    """过滤空值"""
    new_l = []
    for i in old_l:
        if i:
            new_l.append(i)
    return new_l


def data_by_excel(file):
    """
    从excel中加载用例数据
    :param file:
    :return:
    """
    wb = load_workbook(file)
    print_debug_log(f"文件{file=},包含了{len(wb.worksheets)}个sheet页")
    suite_dict = {}  # 以套件名称为key,以用例为value
    for ws in wb.worksheets:
        case_dict = {}  # 以名称为key,以步骤为value的字典
        case_name = ""
        for line in ws.iter_rows(values_only=True):
            _id = line[0]
            step_id.append(line[0])

            print_debug_log(f"正在处理下一行：{line}")
            if line[2] == 'alert' and line[0] is not None:
                settings.cap_png = 0  # 含有alert的页面无法截图,关闭截图功能
            if line[2] == 'img_verify' and line[0] is not None:
                settings.cap_png = 0  # 含有alert的页面无法截图,关闭截图功能
            if line[2] == 'verify_code_gap' and line[0] is not None:
                settings.window_max = 1  # 含有缺口验证码的页面需要最大化
            if isinstance(_id, int):  # 步骤
                if _id == -1:
                    case_name = line[3]
                    case_dict[case_name] = []  # 以用例名称为Key，创建新的空用例
                elif _id > 0:  # 用例名称
                    case_dict[case_name].append(filter_empty(line))
        print_debug_log(f"Sheet'{ws.title}',包含了{len(case_dict)}条用例")
        suite_dict[ws.title] = case_dict
    print_debug_log(f"生成测试用例{suite_dict=}")
    return suite_dict


def create_case(test_suite: dict, file):
    """
    接收从excel而来的多个测试套件的信息，并生成真正的测试用例
    :return:
    """
    file_path = Path(file)
    filename = file_path.name

    def collect_step_id():  # 收集步骤id
        list1 = []
        for i in step_id:
            try:
                a = int(i)
                list1.append(a)
            except:
                pass
            for j in list1:
                if j == -1:
                    list1.remove(j)
        return list1

    gen_id = (i for i in collect_step_id())  # 将步骤id加入生成器

    for suite_name, case_dict in test_suite.items():
        sheet_name.append(suite_name)  # 将sheet页名存入列表
        gen = (name for name in sheet_name)  # 产生一个生成器

        @allure.suite(filename)
        class Test:
            @pytest.fixture(autouse=True)
            def init_pytest(self, request):
                self.request = request
                #  把pytest的夹具保存到测试类当中

            @pytest.mark.parametrize('case', case_dict.items(), ids=case_dict.keys())
            def test_(self, case):
                case_name = (case[0])
                step_list = case[1]
                kw = KeyWord(request=self.request)  # 不传递driver，传递pytest
                sheet_name_ = next(gen)  # 在生成器中取sheet页名

                print_warning_log(f"-----'{sheet_name_}'Sheet页:'{case_name}'用例测试开始-----")
                try:
                    for step in step_list:
                        key = step[2]  # 关键字
                        if key == 'alert':
                            settings.cap_png = 0
                        args = step[3:]  # 关键字参数
                        new_step_id = next(gen_id)  # 从步骤id生成器中取值
                        print_info_log(f"执行第{new_step_id}步关键字：{key=},{args=}")
                        f = kw.get_kw_method(key)  # 调用关键字

                        try:
                            with allure.step(step[1]):
                                f(*args)  # 用例是在这里执行的
                                if settings.cap_png:
                                    allure.attach(
                                        kw.driver.get_screenshot_as_png(),
                                        step[1],
                                        allure.attachment_type.PNG,
                                    )
                        except Exception as e:
                            print_error_log(f'关键字"{key=}"调用出错')
                            # 执行关键字之后截图
                            allure.attach(
                                kw.driver.get_screenshot_as_png(),
                                step[1],
                                allure.attachment_type.PNG,
                            )
                            raise e
                        finally:
                            if settings.cap_png:
                                allure.attach(
                                    kw.driver.get_screenshot_as_png(),
                                    step[1],
                                    allure.attachment_type.PNG,
                                )
                        print_info_log(f"执行关键字：{key=}成功")
                    print_warning_log(f"-----'{sheet_name_}'Sheet页:'{case_name}'用例测试结束-----")
                except Exception as e:
                    print_error_log(f'{suite_name}--{case_name}测试失败')
                    raise e

        print_debug_log(f"生成了测试用例{suite_name}")
        yield Test, suite_name  # 返回了用例和套件名
